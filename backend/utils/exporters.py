"""M8 报表导出：Excel 班级月报 + 个人学期 PDF + 单日班级 Excel。

依赖：openpyxl, reportlab（已在 requirements.txt）
"""
from __future__ import annotations

import io
from datetime import date, timedelta
from typing import List

from django.db.models import Avg, Count, Q


# ────────────────────────────────────────────────────────────────
# Excel 班级月报
# ────────────────────────────────────────────────────────────────

def export_class_month_excel(classroom, year: int, month: int) -> bytes:
    """生成班级指定月份的 Excel 报表，返回字节流。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from apps.sleep.models import SleepRecord

    # 计算月份日期范围
    from calendar import monthrange
    _, last_day = monthrange(year, month)
    start = date(year, month, 1)
    end = date(year, month, last_day)
    dates = [start + timedelta(days=i) for i in range((end - start).days + 1)]

    students = list(classroom.students.select_related("user").order_by("student_no"))

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    # 标题行
    ws.merge_cells(f"A1:{get_column_letter(len(dates) + 3)}1")
    ws["A1"] = f"{classroom.name} · {year}年{month:02d}月 睡眠月报"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # 列头：学号 | 姓名 | 平均分 | 各日期…
    headers = ["学号", "姓名", "月均质量分"] + [d.strftime("%m/%d") for d in dates]
    ws.append([""] + [""] * (len(dates) + 2))  # 空行
    ws.append(headers)

    header_row = ws.max_row
    fill_blue = PatternFill("solid", fgColor="4472C4")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = fill_blue
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 状态颜色
    status_fill = {
        "normal":   PatternFill("solid", fgColor="C6EFCE"),
        "warning":  PatternFill("solid", fgColor="FFEB9C"),
        "abnormal": PatternFill("solid", fgColor="FFC7CE"),
        "severe":   PatternFill("solid", fgColor="FF0000"),
        "missed":   PatternFill("solid", fgColor="F2F2F2"),
    }

    for student in students:
        records = {
            r.date: r
            for r in SleepRecord.objects.filter(student=student, date__range=(start, end))
        }
        valid = [r for r in records.values() if r.status != "missed"]
        avg_q = round(sum(r.quality_score for r in valid) / len(valid), 1) if valid else 0

        row = [student.student_no, student.real_name, avg_q]
        for d in dates:
            r = records.get(d)
            if r and r.status != "missed":
                row.append(r.quality_score)
            else:
                row.append("")
        ws.append(row)

        # 给日期列染色
        data_row = ws.max_row
        for col_idx, d in enumerate(dates, start=4):
            r = records.get(d)
            if r:
                cell = ws.cell(row=data_row, column=col_idx)
                cell.fill = status_fill.get(r.status, PatternFill())

    # 自适应列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(8, min(max_len + 2, 20))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ────────────────────────────────────────────────────────────────
# 单日班级 Excel
# ────────────────────────────────────────────────────────────────

def export_day_overview_excel(classroom, target_date: date) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from apps.sleep.models import SleepRecord

    students = list(classroom.students.select_related("user").order_by("student_no"))

    wb = Workbook()
    ws = wb.active
    ws.title = target_date.isoformat()
    ws["A1"] = f"{classroom.name} · {target_date} 单日概览"
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["学号", "姓名", "入睡时间", "起床时间", "时长(min)", "质量分", "状态", "心情"])

    for student in students:
        r = SleepRecord.objects.filter(student=student, date=target_date).first()
        if r and r.status != "missed":
            ws.append([
                student.student_no,
                student.real_name,
                r.bedtime.strftime("%H:%M") if r.bedtime else "—",
                r.wake_time.strftime("%H:%M") if r.wake_time else "—",
                r.duration_minutes,
                r.quality_score,
                r.get_status_display(),
                r.get_mood_tag_display() if r.mood_tag else "—",
            ])
        else:
            ws.append([student.student_no, student.real_name, "—", "—", 0, 0, "未打卡", "—"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ────────────────────────────────────────────────────────────────
# 个人学期 PDF
# ────────────────────────────────────────────────────────────────

def _register_reportlab_cjk_font() -> str:
    """Return a ReportLab font name that can render Chinese text."""
    import os

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfbase.ttfonts import TTFont

    font_paths = [
        # Windows
        r"C:\Windows\Fonts\msyh.ttc",
        r"C:\Windows\Fonts\simhei.ttf",
        r"C:\Windows\Fonts\simsun.ttc",
        # Ubuntu / Debian
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSerifCJK-Regular.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/arphic/uming.ttc",
        # macOS
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Light.ttc",
    ]

    for fp in font_paths:
        if not os.path.exists(fp):
            continue
        try:
            pdfmetrics.registerFont(TTFont("SleepSystemCJK", fp))
            return "SleepSystemCJK"
        except Exception:
            continue

    # Built into ReportLab. This keeps PDFs readable even before OS fonts are installed.
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        return "STSong-Light"
    except Exception:
        return "Helvetica"


def export_student_semester_pdf(student, year: int, semester: int) -> bytes:
    """semester=1 → 春（3-8月），semester=2 → 秋（9-次年2月）。"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from apps.sleep.models import SleepRecord

    cn_font = _register_reportlab_cjk_font()

    if semester == 1:
        start, end = date(year, 3, 1), date(year, 8, 31)
    else:
        start, end = date(year, 9, 1), date(year + 1, 2, 28)

    records = list(SleepRecord.objects.filter(student=student, date__range=(start, end)).order_by("date"))
    valid = [r for r in records if r.status != "missed"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    cn_normal = ParagraphStyle("cn_normal", fontName=cn_font, fontSize=10, leading=16)
    cn_title = ParagraphStyle("cn_title", fontName=cn_font, fontSize=16, leading=22, spaceAfter=6)
    cn_h2 = ParagraphStyle("cn_h2", fontName=cn_font, fontSize=12, leading=18, spaceBefore=10)

    story = []
    story.append(Paragraph(f"睡眠健康学期报告", cn_title))
    story.append(Paragraph(
        f"学生：{student.real_name}（{student.student_no}）　"
        f"学期：{year}年{'上' if semester == 1 else '下'}学期　"
        f"统计区间：{start} 至 {end}",
        cn_normal
    ))
    story.append(Spacer(1, 6*mm))

    # 统计摘要
    avg_q = round(sum(r.quality_score for r in valid) / len(valid), 1) if valid else 0
    avg_dur = round(sum(r.duration_minutes for r in valid) / len(valid)) if valid else 0
    checked = len(valid)
    total_days = (end - start).days + 1
    status_counts = {s: sum(1 for r in records if r.status == s) for s in ["normal","warning","abnormal","severe","missed"]}

    story.append(Paragraph("一、睡眠数据摘要", cn_h2))
    summary_data = [
        ["指标", "数值"],
        ["打卡天数", f"{checked} / {total_days} 天"],
        ["平均质量分", str(avg_q)],
        ["平均睡眠时长", f"{avg_dur // 60}h {avg_dur % 60}min"],
        ["健康天数", str(status_counts.get("normal", 0))],
        ["预警天数", str(status_counts.get("warning", 0) + status_counts.get("abnormal", 0))],
        ["严重异常天数", str(status_counts.get("severe", 0))],
        ["未打卡天数", str(status_counts.get("missed", 0))],
    ]
    t = Table(summary_data, colWidths=[80*mm, 60*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), cn_font),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
    ]))
    story.append(t)
    story.append(Spacer(1, 6*mm))

    # 明细表（最近 30 条）
    story.append(Paragraph("二、近期打卡明细（最近 30 条）", cn_h2))
    detail_header = ["日期", "入睡", "起床", "时长", "质量分", "状态"]
    detail_rows = [detail_header]
    for r in records[-30:]:
        detail_rows.append([
            str(r.date),
            r.bedtime.strftime("%H:%M") if r.bedtime else "—",
            r.wake_time.strftime("%H:%M") if r.wake_time else "—",
            f"{r.duration_minutes // 60}h{r.duration_minutes % 60}m" if r.duration_minutes else "—",
            str(r.quality_score),
            r.get_status_display(),
        ])

    t2 = Table(detail_rows, colWidths=[30*mm, 25*mm, 25*mm, 25*mm, 20*mm, 25*mm])
    status_color_map = {
        "健康": colors.HexColor("#C6EFCE"),
        "一般": colors.HexColor("#FFEB9C"),
        "异常": colors.HexColor("#FFC7CE"),
        "严重": colors.HexColor("#FF7070"),
        "未打卡": colors.HexColor("#F2F2F2"),
    }
    row_colors = []
    for i, row in enumerate(detail_rows[1:], start=1):
        status_label = row[5]
        fill_color = status_color_map.get(status_label, colors.white)
        row_colors.append(("BACKGROUND", (0, i), (-1, i), fill_color))

    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), cn_font),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        *row_colors,
    ]))
    story.append(t2)

    doc.build(story)
    return buf.getvalue()
